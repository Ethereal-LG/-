from openpyxl import load_workbook
import datetime
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

pd.set_option('display.max_columns', None)   #显示完整的列
pd.set_option('display.max_rows', None)  #显示完整的行

wb = load_workbook('E:\\附件1-yellow_tripdata_2019-01-new.xlsx')
data = pd.read_excel('E:\\附件1-yellow_tripdata_2019-01-new.xlsx')

d = data['PULocationID'].value_counts()
#print(d)

sh = wb["yellow_tripdata_2019-01-new"]

#date = ['2019-01-18']
#date = ['2019-01-02','2019-01-04','2019-01-07','2019-01-10','2019-01-11','2019-01-14','2019-01-15','2019-01-16','2019-01-17','2019-01-19','2019-01-21','2019-01-22','2019-01-23','2019-01-26','2019-01-27','2019-01-28','2019-01-31']
weekend = ['2019-01-05','2019-01-06','2019-01-12','2019-01-13','2019-01-19','2019-01-20','2019-01-26','2019-01-27']

a2 = {}
num = [0]*31
con = '2019-01-01'
num = [0]*31

def PULocationID_count():
    for i in range(1,len(data)):
        lo = data.loc[i,'tpep_pickup_datetime']
        a1 = str(lo).split(' ')[0]
        if a1 in weekend:
            a2[lo] = data.loc[i,'PULocationID']
        count = int(str(a1).split('-')[2])
        num[count-1] += 1
    a3 = pd.Series(a2)
    d0 = a3.value_counts()
    print(d0)
    d0.plot(kind = 'bar')
    plt.show()



def yellow_taxi_orders():
    print(num)
    day = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31]
    plt.plot(day, num, '-ro')
    for i in range(len(num)):
        plt.text(day[i], num[i] + 0.5, '%s' %round(num[i],3), ha='center', fontsize=8)
    plt.xlabel("day") #X轴标签
    plt.ylabel("Yellow taxi orders") #Y轴标签
    plt.show()


a = []
def time_and_PULocationID():
    for i in range(2,sh.max_row+1):
        a.append([])
        for j in range(1,2):
            pick_up_time = sh.cell(i,2).value.strftime('%Y-%m-%d %H-%M-%S')
            area = sh.cell(i,8).value
            a[i-2].append(pick_up_time)
            a[i-2].append(area)
    for i in range(0,sh.max_row-1):
        print(a[i])

propmt = "1.某日各地区订单量统计"
propmt += "\n2.黄色出租车各日总订单量"
propmt += "\n3.各订单请求发出时间和所在地区明细"
propmt += "\n0.退出程序"
while 1:
    print(propmt)
    enter = int(input("请输入您要选择的操作： "))

    if enter == 1:
        PULocationID_count()
    elif enter == 2:
        yellow_taxi_orders()
    elif enter == 3:
        time_and_PULocationID()
    elif enter == 0:
        break
    else:
        print("请重新输入")